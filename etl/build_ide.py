#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Enriquece public/data/indicadores-distrito.json con acceso a servicios básicos
del Censo 2017 (INEI), asistencia escolar (educación) y médicos por 10 000 hab.
(salud), y reconstruye el Índice de Densidad del Estado (IDE), metodología PNUD.

FUENTES (reales, oficiales, descargables):

  1) INEI — REDATAM en línea, base CPV2017DI (Censos Nacionales 2017 a nivel
     distrital). Motor RpWebStats.exe/Frequency. Cada consulta "Frecuencia de la
     variable X con corte por Distrito" devuelve, en una sola pasada, un bloque por
     cada uno de los ~1874 distritos con su UBIGEO de 6 dígitos ("AREA # 010101"),
     los casos por categoría y el total. De ahí calculamos el % correspondiente.
       Vivienda.C2P6   Abastecimiento de agua           -> % agua por red pública
       Vivienda.C2P10  Servicio higiénico (desagüe)     -> % desagüe por red pública
       Vivienda.C2P11  Alumbrado eléctrico por red      -> % electricidad
       Hogar.C3P213    Conexión a Internet              -> % hogares con internet
       Persona (asistencia escolar) con UNIVERSE edad 12–16 -> tasa de asistencia
         escolar de la población de 12 a 16 años (dimensión EDUCACIÓN del IDE).

  2) MINSA — DIGEP / Observatorio de RR.HH. en Salud, base INFORHUS (Registro
     Nacional del Personal de la Salud). Reporte mensual descargable (xlsx) con un
     registro por trabajador y el UBIGEO del ESTABLECIMIENTO donde labora. Contamos
     "MEDICO CIRUJANO" por distrito de trabajo -> médicos por 10 000 hab. (dimensión
     SALUD del IDE, REFERENCIAL: cubre solo el sector MINSA/Gobiernos Regionales —no
     EsSalud/privado/FF.AA.— y es sensible a la movilidad y a la concentración
     hospitalaria: un distrito con un gran hospital "acumula" médicos que atienden a
     toda la provincia).
       Descarga: https://digep.minsa.gob.pe/bdatos.html (enlaces files.minsa.gob.pe).

IDE (reconstrucción QHAWAY, PNUD): promedio simple 0..1 de las dimensiones
DISPONIBLES por distrito. Cada dimensión en 0..1:
  - agua + saneamiento -> ideAgua = promedio(agua%, desague%) / 100
  - electrificación     -> ideElectricidad = electricidad% / 100
  - educación           -> ideEducacion = tasa de asistencia 12–16 / 100
  - salud (referencial) -> ideSalud = min(médicos_10k / TOPE_SALUD, 1)   [min-max 0..tope]
Con las 4 dimensiones cargadas el IDE reconstruido llega a 4/4. calcularIDE() en el
frontend promedia solo las dimensiones presentes (2, 3 o 4).

El script CACHEA en etl/cache_ide/ el HTML crudo de REDATAM y los médicos por ubigeo
para que las re-ejecuciones sean offline y reproducibles. Uso:
    python3 etl/build_ide.py             # usa caché si existe, si no descarga
    python3 etl/build_ide.py --refresh   # fuerza nueva descarga de REDATAM
    python3 etl/build_ide.py --discover  # sonda REDATAM para confirmar variables
"""
import json, os, re, sys, time, html as _html, urllib.request, urllib.parse, ssl, http.cookiejar

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "..", "public", "data")
DEST = os.path.join(DATA, "indicadores-distrito.json")
CACHE = os.path.join(HERE, "cache_ide")
os.makedirs(CACHE, exist_ok=True)

BASE = "CPV2017DI"
HOST = "https://censos2017.inei.gob.pe/bininei"
ACTION = HOST + "/RpWebStats.exe/Frequency?"
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36")

# ctx que ignora verificación TLS (el certificado de INEI a veces falla en curl/py)
_CTX = ssl.create_default_context()
_CTX.check_hostname = False
_CTX.verify_mode = ssl.CERT_NONE

# ── Servicios (Vivienda/Hogar): item REDATAM, universo, tag de salida ──
QUERIES = {
    "agua":         ("Vivienda.C2P6",  "FREQVIV", "Vivienda.TIPOVIV <= 8"),
    "desague":      ("Vivienda.C2P10", "FREQVIV", "Vivienda.TIPOVIV <= 8"),
    "electricidad": ("Vivienda.C2P11", "FREQVIV", "Vivienda.TIPOVIV <= 8"),
    "internet":     ("Hogar.C3P213",   "FREQHOG", ""),
}

# ── Educación (Persona): tasa de asistencia escolar de la población de 12 a 16 años ──
# Aproximación censal de la "tasa neta de asistencia a secundaria" del IDE (PNUD).
# El Censo 2017 preguntó "¿Actualmente asiste a algún colegio/instituto/universidad?"
# (Sí/No), pero NO el nivel al que asiste; por eso medimos ASISTENCIA ESCOLAR en la
# franja etaria 12–16 (edad normativa de secundaria), no la asistencia "a secundaria"
# estricta. Se declara explícitamente como aproximación. Numerador = "Sí" asiste;
# denominador = total de personas 12–16 que respondieron.
#   ROW/ITEM  = variable de asistencia escolar (Persona.C5P12)
#   UNIVERSE  = franja etaria 12..16 sobre la variable de edad (Persona.C5P4)
# Los nombres exactos se confirman en vivo con `--discover` (censos2017 estaba caído
# durante parte del desarrollo). Cambia aquí si REDATAM reporta otros códigos.
EDU_ASIST_ITEM = "Persona.C5P12"
EDU_EDAD_VAR = "Persona.C5P4"
EDU_UNIVERSE = f"{EDU_EDAD_VAR} >= 12 AND {EDU_EDAD_VAR} <= 16"

# candidatos que `--discover` probará a nivel Departamento (rápido) para hallar los
# nombres correctos de las variables de asistencia y de edad.
DISCOVER_ASIST = ["Persona.C5P12", "Persona.C5P14", "Persona.C5P13", "Persona.P08",
                  "Persona.C5AP4", "Persona.ASISTENCIA"]
DISCOVER_EDAD = ["Persona.C5P4", "Persona.EDAD", "Persona.C5P4_R", "Persona.P03"]

# ── Salud: médicos (MINSA/GORE) por 10 000 hab, min-max con tope (referencial) ──
MEDICOS_CACHE = os.path.join(CACHE, "medicos_ubigeo.json")   # {ubigeo: n_medicos}
INFORHUS_XLSX = os.path.join(CACHE, "inforhus.xlsx")         # base mensual DIGEP (opcional)
# Enlace de descarga (comparte Nextcloud tras Cloudflare). Se actualiza cada mes en
# https://digep.minsa.gob.pe/bdatos.html — poner aquí el token del mes vigente.
INFORHUS_SHARE = "https://files.minsa.gob.pe/s/brPb6qFyPGgBm7S"  # BASE MAYO 2026
# Tope de la normalización min-max de médicos/10k (percentil 95 observado ≈ 37).
# Evita que la concentración hospitalaria (p. ej. Arequipa cercado ≈ 198/10k) sature
# la escala. Se recalcula/repite en tiempo de build y se imprime.
SALUD_TOPE_DEFAULT = 37.0


def _opener():
    cj = http.cookiejar.CookieJar()
    return urllib.request.build_opener(
        urllib.request.HTTPSHandler(context=_CTX),
        urllib.request.HTTPCookieProcessor(cj),
    )


def _post(url, data, referer, timeout=240):
    body = urllib.parse.urlencode(data).encode("utf-8")
    req = urllib.request.Request(url, data=body, headers={
        "User-Agent": UA, "Referer": referer,
        "Content-Type": "application/x-www-form-urlencoded",
    })
    with urllib.request.urlopen(req, timeout=timeout, context=_CTX) as r:
        return r.read().decode("utf-8", "ignore")


def _get(url, timeout=240):
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=timeout, context=_CTX) as r:
        return r.read().decode("latin-1", "ignore")


def _redatam_run(row, item, universe, timeout=240):
    """Ejecuta una consulta Frequency (corte por Distrito) y devuelve el HTML crudo
    con los bloques 'AREA #'. Levanta excepción si REDATAM no responde."""
    referer = f"{HOST}/RpWebStats.exe/Frequency?BASE={BASE}&ITEM={item}&lang=esp"
    data = {
        "MAIN": "WebServerMain.inl", "BASE": BASE, "LANG": "esp",
        "CODIGO": "XXUSUARIOXX", "ITEM": item, "MODE": "RUN", "inputTitle": "",
        "ROW": row, "AREABREAK": "Distrito", "SELECTION": "ALL",
        "FORMAT": "HTML", "PERCENT": "OFF", "UNIVERSE": universe,
        "FILTER": "", "TEXT_FILTER": "", "INLINESELECTION": "", "Submit": "Ejecutar",
    }
    shell = _post(ACTION, data, referer, timeout=timeout)
    m = re.search(r"RpWebUtilities\.exe/Text\?LFN=RpBases\\Tempo\\[0-9]+\\~tmp_[0-9]+\.htm", shell)
    if not m:
        raise RuntimeError("REDATAM no devolvió enlace de resultados")
    grid = _get(HOST + "/" + m.group(0) + "&TYPE=TMP", timeout=timeout)
    if "AREA #" not in grid:
        raise RuntimeError("tabla de resultados sin bloques AREA")
    return grid


def fetch_raw(tag, row, item, universe, refresh=False):
    """HTML crudo de la tabla de frecuencia por distrito (con caché en cache_ide/)."""
    cache_file = os.path.join(CACHE, f"{tag}.html")
    if os.path.exists(cache_file) and not refresh:
        return open(cache_file, encoding="utf-8").read()

    last = None
    for attempt in range(3):
        try:
            grid = _redatam_run(row, item, universe)
            open(cache_file, "w", encoding="utf-8").write(grid)
            return grid
        except Exception as e:  # noqa: BLE001
            last = e
            print(f"    [{tag}] reintento {attempt}: {e!r}", flush=True)
            time.sleep(3)
    raise SystemExit(f"No se pudo descargar {tag} de REDATAM: {last!r}")


_NUM = re.compile(r"[0-9]+")


def _to_int(s):
    s = s.replace(".", "").replace(",", "").replace("\xa0", "").replace(" ", "").strip()
    return int(s) if _NUM.fullmatch(s) else None


def parse_blocks(html):
    """HTML REDATAM -> {ubigeo: {'cats': {label: casos}, 'total': int}}."""
    rows = re.split(r"</tr>", html, flags=re.I)
    out, cur = {}, None
    for tr in rows:
        cells = [re.sub(r"<[^>]+>", "", c) for c in re.split(r"</t[dh]>", tr, flags=re.I)]
        cells = [re.sub(r"\s+", " ", _html.unescape(c)).strip() for c in cells]
        cells = [c for c in cells if c != ""]
        if not cells:
            continue
        joined = " ".join(cells)
        m = re.search(r"AREA #\s*([0-9]{6})", joined)
        if m:
            cur = m.group(1)
            out[cur] = {"cats": {}, "total": None}
            continue
        if cur is None or len(cells) < 2:
            continue
        label, val = cells[0], _to_int(cells[1])
        if val is None:
            continue
        if label == "Total":
            out[cur]["total"] = val
        elif not label.startswith("V:") and label not in ("Casos", "No Aplica :"):
            out[cur]["cats"][label] = val
    return out


# ── extractores de "casos con acceso" por variable (match por substring, robusto) ──
def _pct_agua(cats):
    # Definición oficial INEI "agua por red pública": red pública dentro + fuera de la
    # vivienda (78.3% nacional). El pilón/pileta de uso público se contabiliza aparte.
    return sum(v for lab, v in cats.items() if "Red pública" in lab)


def _pct_desague(cats):
    return sum(v for lab, v in cats.items() if "Red pública de desagüe" in lab)


def _pct_si(cats):
    return sum(v for lab, v in cats.items() if lab.lower().startswith("sí") or lab.lower().startswith("si "))


EXTRACTOR = {
    "agua": _pct_agua,
    "desague": _pct_desague,
    "electricidad": _pct_si,
    "internet": _pct_si,
    "educacion": _pct_si,   # asistencia escolar: categoría "Sí" sobre el total 12–16
}


def compute_service(tag, blocks):
    """{ubigeo: pct 0..100} para un servicio/indicador de proporción."""
    fn = EXTRACTOR[tag]
    res = {}
    for ubigeo, b in blocks.items():
        tot = b["total"]
        if not tot:
            continue
        ok = fn(b["cats"])
        res[ubigeo] = round(100.0 * ok / tot, 2)
    return res


def national_pct(tag, blocks):
    num = den = 0
    for b in blocks.values():
        if not b["total"]:
            continue
        num += EXTRACTOR[tag](b["cats"])
        den += b["total"]
    return 100.0 * num / den if den else 0.0


# ───────────────────────── SALUD (médicos/10k) ─────────────────────────

def _download_inforhus():
    """Descarga el xlsx mensual de INFORHUS (DIGEP) sorteando Cloudflare con cookie."""
    op = _opener()
    hdrs = {"User-Agent": UA, "Accept": "text/html,application/xhtml+xml"}
    # 1) visitar la página de compartición para obtener la cookie __cf_bm
    op.open(urllib.request.Request(INFORHUS_SHARE, headers=hdrs), timeout=120).read()
    # 2) descargar el archivo (sigue el 303 al WebDAV)
    req = urllib.request.Request(INFORHUS_SHARE + "/download", headers={"User-Agent": UA, "Accept": "*/*"})
    with op.open(req, timeout=600) as r:
        blob = r.read()
    with open(INFORHUS_XLSX, "wb") as f:
        f.write(blob)
    return INFORHUS_XLSX


def _parse_inforhus_xlsx(path):
    """{ubigeo: n_medicos} contando 'MEDICO CIRUJANO' por UBIGEO del establecimiento."""
    try:
        import openpyxl  # dependencia opcional, solo si hay que reparsear el xlsx
    except ImportError:
        raise SystemExit("Falta openpyxl para parsear INFORHUS (pip install openpyxl)")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["B MAYO"] if "B MAYO" in wb.sheetnames else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    next(it)                    # fila de título ("BASE MAYO 2026")
    header = list(next(it))     # fila de cabeceras reales
    def col(name):
        return header.index(name)
    c_ub = col("UBIGEO"); c_prof = col("profesion")
    med = {}
    for row in it:
        p = row[c_prof]
        if p and str(p).strip().upper() == "MEDICO CIRUJANO":
            ub = row[c_ub]
            if ub is None:
                continue
            ub = str(ub).strip()
            if ub.isdigit():
                ub = ub.zfill(6)
            med[ub] = med.get(ub, 0) + 1
    return med


def load_medicos(refresh=False):
    """{ubigeo: n_medicos}. Prioridad: caché JSON -> xlsx local -> descarga DIGEP."""
    if os.path.exists(MEDICOS_CACHE) and not refresh:
        return {k: int(v) for k, v in json.load(open(MEDICOS_CACHE, encoding="utf-8")).items()}
    path = INFORHUS_XLSX
    if not os.path.exists(path) or refresh:
        print("[MINSA] descargando INFORHUS (DIGEP) …", flush=True)
        try:
            path = _download_inforhus()
        except Exception as e:  # noqa: BLE001
            print(f"    [salud] no se pudo descargar INFORHUS: {e!r}", flush=True)
            return None
    print("[MINSA] parseando INFORHUS (médicos por ubigeo) …", flush=True)
    med = _parse_inforhus_xlsx(path)
    json.dump(med, open(MEDICOS_CACHE, "w", encoding="utf-8"), ensure_ascii=False)
    return med


def r2(x):
    return None if x is None else round(x, 4)


def _percentile(sorted_vals, q):
    import math
    if not sorted_vals:
        return None
    k = (len(sorted_vals) - 1) * q
    f = math.floor(k); c = math.ceil(k)
    if f == c:
        return sorted_vals[f]
    return sorted_vals[f] * (c - k) + sorted_vals[c] * (k - f)


# ───────────────────────── discover (diagnóstico REDATAM) ─────────────────────────

def discover():
    """Sonda candidatos de variables de asistencia/edad a nivel Departamento (rápido)
    e imprime las etiquetas del primer bloque para confirmar los nombres correctos."""
    def probe(item, universe=""):
        referer = f"{HOST}/RpWebStats.exe/Frequency?BASE={BASE}&ITEM={item}&lang=esp"
        data = {
            "MAIN": "WebServerMain.inl", "BASE": BASE, "LANG": "esp",
            "CODIGO": "XXUSUARIOXX", "ITEM": item, "MODE": "RUN", "inputTitle": "",
            "ROW": item, "AREABREAK": "Departamento", "SELECTION": "ALL",
            "FORMAT": "HTML", "PERCENT": "OFF", "UNIVERSE": universe,
            "FILTER": "", "TEXT_FILTER": "", "INLINESELECTION": "", "Submit": "Ejecutar",
        }
        shell = _post(ACTION, data, referer, timeout=120)
        m = re.search(r"RpWebUtilities\.exe/Text\?LFN=RpBases\\Tempo\\[0-9]+\\~tmp_[0-9]+\.htm", shell)
        if not m:
            return None
        grid = _get(HOST + "/" + m.group(0) + "&TYPE=TMP", timeout=120)
        blocks = parse_blocks(grid)
        first = next(iter(blocks.values()), None)
        return first["cats"] if first else {}

    print("── candidatos ASISTENCIA (esperado: Sí/No) ──")
    for it in DISCOVER_ASIST:
        try:
            cats = probe(it)
            print(f"  {it}: {list(cats.keys())[:6] if cats is not None else 'ERROR'}")
        except Exception as e:  # noqa: BLE001
            print(f"  {it}: {e!r}")
    print("── candidatos EDAD (esperado: 0,1,2,… o rangos) ──")
    for it in DISCOVER_EDAD:
        try:
            cats = probe(it)
            print(f"  {it}: {list(cats.keys())[:8] if cats is not None else 'ERROR'}")
        except Exception as e:  # noqa: BLE001
            print(f"  {it}: {e!r}")


# ───────────────────────────────── main ─────────────────────────────────

def main():
    refresh = "--refresh" in sys.argv
    if "--discover" in sys.argv:
        discover()
        return

    # ── 1) Servicios básicos del Censo 2017 (agua, desagüe, luz, internet) ──
    services = {}
    for tag, (item, _uni_kind, uni) in QUERIES.items():
        print(f"[REDATAM] {tag}: {item} por Distrito …", flush=True)
        html = fetch_raw(tag, item, item, uni, refresh=refresh)
        blocks = parse_blocks(html)
        services[tag] = compute_service(tag, blocks)
        print(f"    distritos={len(services[tag])}  nacional≈{national_pct(tag, blocks):.1f}%", flush=True)

    # ── 2) Educación: tasa de asistencia escolar 12–16 (Censo 2017) ──
    educ = {}
    edu_nat = None
    try:
        print(f"[REDATAM] educacion: {EDU_ASIST_ITEM} | universo {EDU_UNIVERSE} …", flush=True)
        html = fetch_raw("educacion", EDU_ASIST_ITEM, EDU_ASIST_ITEM, EDU_UNIVERSE, refresh=refresh)
        blocks = parse_blocks(html)
        educ = compute_service("educacion", blocks)
        edu_nat = national_pct("educacion", blocks)
        print(f"    distritos={len(educ)}  asistencia 12–16 nacional≈{edu_nat:.1f}%", flush=True)
    except SystemExit as e:
        print(f"    [educacion] no disponible ({e}); ideEducacion quedará null.", flush=True)

    # ── 3) Salud: médicos (MINSA/GORE) por 10 000 hab (INFORHUS) ──
    medicos = load_medicos(refresh=refresh)

    # ── merge con el JSON existente ──
    base = json.load(open(DEST, encoding="utf-8"))

    # tope de la normalización min-max de salud (p95 observado; se imprime)
    salud_tope = SALUD_TOPE_DEFAULT
    if medicos:
        pob_by_ub = {d["ubigeo"]: d.get("pob") for d in base}
        dens_vals = sorted(
            10000.0 * medicos[ub] / pob_by_ub[ub]
            for ub in medicos
            if pob_by_ub.get(ub) and pob_by_ub[ub] > 0
        )
        p95 = _percentile(dens_vals, 0.95)
        if p95:
            salud_tope = round(p95, 1)

    n_serv = n_edu = n_sal = 0
    for d in base:
        ub = d["ubigeo"]
        agua = services["agua"].get(ub)
        des = services["desague"].get(ub)
        ele = services["electricidad"].get(ub)
        net = services["internet"].get(ub)
        # servicios (0..100)
        d["agua"] = agua
        d["desague"] = des
        d["electricidad"] = ele
        d["internet"] = net
        # dimensión IDE agua+saneamiento (promedio de agua y desagüe)
        if agua is not None and des is not None:
            d["ideAgua"] = r2((agua + des) / 2.0 / 100.0)
        elif agua is not None:
            d["ideAgua"] = r2(agua / 100.0)
        else:
            d["ideAgua"] = None
        d["ideElectricidad"] = r2(ele / 100.0) if ele is not None else None

        # dimensión educación: asistencia escolar 12–16 como proporción 0..1
        edu = educ.get(ub)
        d["ideEducacion"] = r2(edu / 100.0) if edu is not None else None

        # dimensión salud (referencial): min-max de médicos/10k con tope
        if medicos is not None and isinstance(d.get("pob"), (int, float)) and d["pob"] > 0:
            dens = 10000.0 * medicos.get(ub, 0) / d["pob"]
            d["ideSalud"] = r2(min(dens / salud_tope, 1.0))
        else:
            d["ideSalud"] = None

        if agua is not None:
            n_serv += 1
        if d["ideEducacion"] is not None:
            n_edu += 1
        if d["ideSalud"] is not None:
            n_sal += 1

    json.dump(base, open(DEST, "w", encoding="utf-8"),
              ensure_ascii=False, separators=(",", ":"))

    # dimensiones disponibles a nivel nacional (para el título "N de 4")
    dims_presentes = 2  # agua+saneamiento y electrificación (siempre del Censo 2017)
    if n_edu:
        dims_presentes += 1
    if n_sal:
        dims_presentes += 1

    print(f"\nOK -> {DEST}")
    print(f"    {n_serv}/{len(base)} distritos con servicios del Censo 2017")
    print(f"    educación: {n_edu} distritos con ideEducacion"
          + (f" (asistencia 12–16 nacional≈{edu_nat:.1f}%)" if edu_nat is not None else ""))
    print(f"    salud: {n_sal} distritos con ideSalud (tope min-max = {salud_tope}/10k)")
    print(f"    IDE reconstruido a {dims_presentes} de 4 dimensiones")


if __name__ == "__main__":
    main()
